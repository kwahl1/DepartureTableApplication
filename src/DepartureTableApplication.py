'''

Written and maintained by:
Kristian Wahlqvist
kristian.wahlqvist@afry.com
2019-09

------------------------------------------------------------------------------------

The goal of this project was to create a program that can show SL's real-time 
departure information of stations nearby the Solna office. The information 
should be displayed as part of the Powerpoint presentation rolling on the TV 
screen on floor 7. This way, employes that are passing by can conviniently see 
when the next bus or train is leaving.

See the full Wiki on this project at:
https://teams.microsoft.com/l/channel/19%3A03dd9d6bdea14660ad2c2c651f0c738e%40thread.skype/tab%3A%3A531e917d-b1bd-4a01-b620-6a16439a0649?groupId=9475dc85-3d0c-4534-82bd-0e815ee72a72&tenantId=58af3eba-510e-4544-8cfd-85f5e0206382

------------------------------------------------------------------------------------


'''

from openpyxl import load_workbook
from datetime import datetime
import time
import os, os.path
import win32com.client
import urllib.request
import json
import time
from datetime import datetime
import yaml as yaml
import io
import tkinter as tk
import random


class DepartureTableApplication():
	def __init__(self):
		# Constructor
		settings = yaml.safe_load(open("C:\\User\\Name\\Documents\\DepartureTableApplication\\config\\user_parameters.yaml",'r',encoding = 'utf-8'))
		# get parameters
		self.TIME_WINDOWS = []
		self.KEY_ID = settings["key_id"]
		self.TIME_WINDOWS.append(settings["realtidsinfo"]["time_window_bus"])
		self.TIME_WINDOWS.append(settings["realtidsinfo"]["time_window_train"])
		self.UPDATE_FREQUENCY = int(settings["realtidsinfo"]["update_time"])
		self.STATION_ID = settings["site_id"]
		self.WORKING_DIR = settings["working_directory"]
		self.MAX_ROWS = settings["max_rows"]

		self.PPT_SL = None
		self.PPT_SLIDESHOW = None
		self.PPT_APP = None

		if os.path.exists(self.WORKING_DIR + "\\src\\sl_realtid_ppt.pptm"): 
			self.PPT_APP=win32com.client.Dispatch("Powerpoint.Application")
			# slide with departure times and relevant VBA macros
			self.PPT_SL = self.PPT_APP.Presentations.Open(self.WORKING_DIR + "\\src\\sl_realtid_ppt.pptm", ReadOnly=1)
			# merge with target powerpoint in folder /Powerpoint
			self.PPT_SL.Application.Run("sl_realtid_ppt.pptm!Module4.mergePresentations")
			self.PPT_SLIDESHOW = self.PPT_APP.Presentations.Open(self.WORKING_DIR + "\\src\\slideshow.pptm", ReadOnly=1)
			# start slideshow
			self.PPT_SLIDESHOW.Application.Run("sl_realtid_ppt.pptm!Module3.slideshow")
		else:
			raise FileNotFoundError("\nFile not found: " + self.WORKING_DIR + "\\src\\sl_realtid_ppt.pptm" 
				+ "\nTry modifying working_directory in config/user_parameters.yaml\n")

	def __del__(self):
		# Destructor
		print("Quit.")
		if(self.PPT_SL != None):
			self.PPT_SL.close()
		if(self.PPT_SLIDESHOW != None):
			self.PPT_SLIDESHOW.close()
		if(self.PPT_APP != None):
			self.PPT_APP.Quit()

	def getDepartures(self, key,siteID,time_window):
		response = urllib.request.urlopen("http://api.sl.se/api2/realtimedeparturesv4.json?key="+key+"&siteid="+siteID+"&timewindow="+time_window).read()
		parsed = json.loads(response)

		while(parsed["StatusCode"] != 0):

			self.log("No response from URL request. Retrying..")
			time.sleep(0.5)
			response = urllib.request.urlopen("http://api.sl.se/api2/realtimedeparturesv4.json?key="+key+"&siteid="+siteID+"&timewindow="+time_window).read()
			parsed = json.loads(response)

		return parsed["ResponseData"]

	def printTimeTable(self, station, transport_type, data):
		# for terminal output
		print("\n\t--- "+station+" ---")
		print("{:10}{:25}{:10}".format("LINE", "DESTINATION", "DEPARTURE"))

		for departure in data[transport_type]:
			print('{:10}{:25}{:10}'.format(departure["LineNumber"], departure["Destination"], departure["DisplayTime"]))

		print("\n")

	def log(self, msg):
		# TODO write to file
		print(self.getTimestampNow()+": "+msg)

	def getTimestampNow(self):
		# current time
		now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
		return now_str

	def writeToExcell(self):
		sl_data_kolonn = self.getDepartures(self.KEY_ID, self.STATION_ID["Kolonnvägen"], self.TIME_WINDOWS[0])
		sl_data_solna = self.getDepartures(self.KEY_ID, self.STATION_ID["Solna Station"], self.TIME_WINDOWS[1])

		wb = load_workbook(self.WORKING_DIR+'\\src\\sl_realtid_excell.xlsx',read_only = False, data_only=True)
		ws = wb['SL']

		self.clearSheet(ws)

		# e.g. set value in cell 'A3'
		for row, departure in enumerate(sl_data_kolonn["Buses"],4):

			ws['A'+str(row)] = departure["LineNumber"]
			ws['B'+str(row)] = departure["Destination"]
			ws['C'+str(row)] = departure["DisplayTime"]
			if row >= int(self.MAX_ROWS): break

		for row, departure in enumerate(sl_data_solna["Trains"],4):

			ws['E'+str(row)] = departure["LineNumber"]
			ws['F'+str(row)] = departure["Destination"]
			ws['G'+str(row)] = departure["DisplayTime"]
			if row >= int(self.MAX_ROWS): break

		for row, departure in enumerate(sl_data_solna["Buses"],4):

			ws['H'+str(row)] = departure["LineNumber"]
			ws['I'+str(row)] = departure["Destination"]
			ws['J'+str(row)] = departure["DisplayTime"]
			if row >= int(self.MAX_ROWS): break

		for row, departure in enumerate(sl_data_solna["Trams"],4):

			ws['K'+str(row)] = departure["LineNumber"]
			ws['L'+str(row)] = departure["Destination"]
			ws['M'+str(row)] = departure["DisplayTime"]
			if row >= int(self.MAX_ROWS): break

		ws["A"+str(int(self.MAX_ROWS)+1)] = "Uppdaterad: "+self.getTimestampNow()

		try:

			wb.save(self.WORKING_DIR+'\\src\\sl_realtid_excell.xlsx')
			self.log("Updated")

		except PermissionError:
			self.log("PermissionError. Skipping update.")

		wb.close()

	def clearSheet(self, ws):
			#clear previous data from sheet
			for row in range(4,int(self.MAX_ROWS)+1):

				ws['A'+str(row)] = ""
				ws['B'+str(row)] = ""
				ws['C'+str(row)] = ""
				ws['E'+str(row)] = ""
				ws['F'+str(row)] = ""
				ws['G'+str(row)] = ""
				ws['H'+str(row)] = ""
				ws['I'+str(row)] = ""
				ws['J'+str(row)] = ""
				ws['K'+str(row)] = ""
				ws['L'+str(row)] = ""
				ws['M'+str(row)] = ""

class HelperGUI(tk.Frame):
	# Simple GUI to end the process after button press
    def __init__(self, master, departureObj):
        super().__init__(master)
        self.master = master
        self.departureObj = departureObj
       	self.master.title("Quit SL")
        self.pack(padx = 50, pady = 50)
        self.create_button()

        self.loop()

    def create_button(self):
        self.quit = tk.Button(self, text="QUIT",font="Arial 16 bold", fg="black", command=self.stop)
        self.quit.config(width = 10)
        self.quit.pack(side="top")

    def stop(self):
    	del self.departureObj
    	self.master.quit()

    def loop(self):
    	self.departureObj.writeToExcell()
    	# add random time to avoid clashing file access with ppt
    	T = self.departureObj.UPDATE_FREQUENCY + random.randint(-2,2)
    	self.after(1000*T,self.loop) #to milliseconds


def main():
	try:
		departureObj = DepartureTableApplication()
		root = tk.Tk()
		app = HelperGUI(root, departureObj)
		app.mainloop()
	except Exception as err:
		print(str(err))


main()


